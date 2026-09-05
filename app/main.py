import os
import base64
import cv2
import numpy as np
from datetime import datetime, date
from typing import Optional
from fastapi import FastAPI, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".env"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = FastAPI(title="Focus Logger")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
EXCEL_FILE = os.path.join(BASE_DIR, os.getenv("EXCEL_FILE", "focus_log.xlsx"))
FACE_CASCADE_PATH = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
CHECKIN_INTERVAL_MIN = int(os.getenv("CHECKIN_INTERVAL_MIN", "22"))
face_cascade = cv2.CascadeClassifier(FACE_CASCADE_PATH)
class ClockInRequest(BaseModel):
    frame: str
class SessionStartRequest(BaseModel):
    type: str
    note: str
class CheckinRequest(BaseModel):
    session_id: int
    status: str
    note: str
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws_sessions = wb.active
        ws_sessions.title = "Sessions"
        ws_sessions.append(["session_id", "date", "start_time", "end_time", "duration_min", "type", "note"])
        ws_checkins = wb.create_sheet("Checkins")
        ws_checkins.append(["checkin_id", "session_id", "timestamp", "status", "note"])
        wb.save(EXCEL_FILE)
def get_next_session_id():
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Sessions"]
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0] > max_id:
            max_id = row[0]
    return max_id + 1
def get_next_checkin_id():
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Checkins"]
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0] > max_id:
            max_id = row[0]
    return max_id + 1
def get_open_session():
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Sessions"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] is None:
            return {
                "session_id": row[0],
                "date": row[1],
                "start_time": row[2],
                "end_time": row[3],
                "duration_min": row[4],
                "type": row[5],
                "note": row[6],
            }
    return None
def close_open_session():
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Sessions"]
    for row in ws.iter_rows(min_row=2):
        if row[3].value is None:
            end_time = datetime.now()
            start_time = row[2].value
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time)
            duration = (end_time - start_time).total_seconds() / 60
            row[3].value = end_time
            row[4].value = round(duration, 2)
            wb.save(EXCEL_FILE)
            return row[0].value
    return None
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})
@app.post("/clock-in")
async def clock_in(req: ClockInRequest):
    try:
        img_data = base64.b64decode(req.frame.split(",")[1])
        nparr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(30, 30))
        return {"face_detected": len(faces) > 0}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
@app.post("/session/start")
async def start_session(req: SessionStartRequest):
    init_excel()
    close_open_session()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Sessions"]
    session_id = get_next_session_id()
    now = datetime.now()
    ws.append([session_id, now.date(), now, None, None, req.type, req.note])
    wb.save(EXCEL_FILE)
    return {"session_id": session_id, "start_time": now.isoformat()}
@app.post("/session/checkin")
async def log_checkin(req: CheckinRequest):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Checkins"]
    checkin_id = get_next_checkin_id()
    now = datetime.now()
    ws.append([checkin_id, req.session_id, now, req.status, req.note])
    wb.save(EXCEL_FILE)
    return {"checkin_id": checkin_id, "timestamp": now.isoformat()}
@app.post("/session/end-day")
async def end_day():
    closed_id = close_open_session()
    return {"closed_session_id": closed_id}
@app.get("/sessions/today")
async def get_today_sessions():
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Sessions"]
    today = date.today()
    sessions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date = row[1].date() if hasattr(row[1], 'date') else row[1]
        if row_date == today:
            sessions.append({
                "session_id": row[0],
                "date": str(row_date),
                "start_time": row[2].isoformat() if row[2] else None,
                "end_time": row[3].isoformat() if row[3] else None,
                "duration_min": row[4],
                "type": row[5],
                "note": row[6],
            })
    return {"sessions": sessions}
@app.get("/session/current")
async def get_current_session():
    session = get_open_session()
    if session:
        session["start_time"] = session["start_time"].isoformat() if session["start_time"] else None
        session["end_time"] = session["end_time"].isoformat() if session["end_time"] else None
    return {"session": session}
@app.get("/checkin/interval")
async def get_checkin_interval():
    return {"interval_min": CHECKIN_INTERVAL_MIN}
@app.get("/session/{session_id}/checkins")
async def get_session_checkins(session_id: int):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Checkins"]
    checkins = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == session_id:
            checkins.append({
                "checkin_id": row[0],
                "session_id": row[1],
                "timestamp": row[2].isoformat() if row[2] else None,
                "status": row[3],
                "note": row[4],
            })
    return {"checkins": checkins}
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="localhost", port=8000)